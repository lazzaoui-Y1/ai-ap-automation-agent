"""
backend/app/erp_connectors/excel_connector.py
موصل Excel - حفظ الفواتير في ملفات Excel
"""

from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .base_connector import BaseERPConnector, ConnectorType, ConnectorStatus
from ..schemas.invoice_schema import Invoice
from ..utils.exceptions import ERPConnectionError, ERPDataFormatError


class ExcelConnector(BaseERPConnector):
    """
    موصل Excel - يحفظ الفواتير في ملف Excel
    
    الإعدادات المطلوبة في config:
        output_path: مسار ملف Excel
        sheet_name: اسم الورقة (افتراضي: "Invoices")
        auto_create: إنشاء الملف تلقائياً (افتراضي: True)
    """
    
    def __init__(self, customer_id: str, config: Dict[str, Any], connector_type: ConnectorType):
        super().__init__(customer_id, config, connector_type)
        
        # الإعدادات
        self.output_path = Path(config.get('output_path', f'./customers/{customer_id}/data/invoices.xlsx'))
        self.sheet_name = config.get('sheet_name', 'Invoices')
        self.auto_create = config.get('auto_create', True)
        
        # Workbook
        self.workbook = None
        self.worksheet = None
        
        # Headers
        self.headers = [
            'Invoice Number', 'Date', 'Vendor Name', 'Vendor Tax ID',
            'Subtotal', 'Tax', 'Discount', 'Total Amount', 'Currency',
            'Line Items Count', 'Language', 'Confidence Score',
            'Processed Date', 'Status', 'Notes'
        ]
    
    def connect(self) -> bool:
        """فتح أو إنشاء ملف Excel"""
        try:
            self.status = ConnectorStatus.CONNECTED
            
            # إنشاء المجلد إذا لم يكن موجوداً
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # فتح أو إنشاء الملف
            if self.output_path.exists():
                self.workbook = openpyxl.load_workbook(self.output_path)
                
                # الحصول على الورقة أو إنشاؤها
                if self.sheet_name in self.workbook.sheetnames:
                    self.worksheet = self.workbook[self.sheet_name]
                else:
                    self.worksheet = self.workbook.create_sheet(self.sheet_name)
                    self._create_headers()
            
            elif self.auto_create:
                # إنشاء ملف جديد
                self.workbook = openpyxl.Workbook()
                self.worksheet = self.workbook.active
                self.worksheet.title = self.sheet_name
                self._create_headers()
                self.workbook.save(self.output_path)
            
            else:
                raise ERPConnectionError(
                    f"Excel file not found: {self.output_path}",
                    erp_system="excel"
                )
            
            self._log_operation("connect", "success")
            self.logger.info(f"Excel connector connected: {self.output_path}")
            
            return True
        
        except Exception as e:
            self.status = ConnectorStatus.ERROR
            self.last_error = str(e)
            self._log_operation("connect", "failed", {"error": str(e)})
            raise ERPConnectionError(
                f"Failed to connect to Excel: {str(e)}",
                erp_system="excel"
            )
    
    def disconnect(self) -> bool:
        """حفظ وإغلاق ملف Excel"""
        try:
            if self.workbook:
                self.workbook.save(self.output_path)
                self.workbook.close()
            
            self.status = ConnectorStatus.DISCONNECTED
            self._log_operation("disconnect", "success")
            
            return True
        
        except Exception as e:
            self.logger.error(f"Failed to disconnect Excel: {str(e)}")
            return False
    
    def send_invoice(self, invoice: Invoice) -> Dict[str, Any]:
        """إضافة فاتورة إلى Excel"""
        try:
            # التحقق من الاتصال
            if not self.workbook or not self.worksheet:
                self.connect()
            
            # التحقق من البيانات
            self._validate_invoice_data(invoice)
            
            # التحقق من التكرار
            if self._check_duplicate(invoice.invoice_number):
                self.logger.warning(f"Invoice {invoice.invoice_number} already exists in Excel")
                # يمكن تحديث البيانات بدلاً من رفع استثناء
                return {
                    "success": True,
                    "erp_id": invoice.invoice_number,
                    "message": "Invoice already exists (not added again)",
                    "details": {"action": "skipped"}
                }
            
            # إضافة الصف
            row_data = self._invoice_to_row(invoice)
            self.worksheet.append(row_data)
            
            # تنسيق الصف
            row_num = self.worksheet.max_row
            self._format_row(row_num)
            
            # حفظ
            self.workbook.save(self.output_path)
            
            # تحديث الإحصائيات
            self._update_sync_stats(success=True)
            
            self._log_operation(
                "send_invoice",
                "success",
                {"invoice_number": invoice.invoice_number}
            )
            
            self.logger.info(f"Invoice {invoice.invoice_number} added to Excel")
            
            return {
                "success": True,
                "erp_id": invoice.invoice_number,
                "message": "Invoice added successfully",
                "details": {
                    "file_path": str(self.output_path),
                    "sheet_name": self.sheet_name,
                    "row_number": row_num
                }
            }
        
        except Exception as e:
            self._update_sync_stats(success=False)
            self.last_error = str(e)
            
            self._log_operation(
                "send_invoice",
                "failed",
                {"invoice_number": invoice.invoice_number, "error": str(e)}
            )
            
            raise ERPDataFormatError(
                f"Failed to add invoice to Excel: {str(e)}"
            )
    
    def test_connection(self) -> bool:
        """اختبار الاتصال"""
        try:
            if self.connect():
                self.disconnect()
                return True
            return False
        except:
            return False
    
    def _create_headers(self):
        """إنشاء عناوين الأعمدة"""
        for col_num, header in enumerate(self.headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            
            # تنسيق العنوان
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # تعيين عرض الأعمدة
        column_widths = {
            'A': 20,  # Invoice Number
            'B': 15,  # Date
            'C': 30,  # Vendor Name
            'D': 20,  # Vendor Tax ID
            'E': 15,  # Subtotal
            'F': 15,  # Tax
            'G': 15,  # Discount
            'H': 15,  # Total
            'I': 10,  # Currency
            'J': 15,  # Line Items
            'K': 12,  # Language
            'L': 15,  # Confidence
            'M': 20,  # Processed Date
            'N': 15,  # Status
            'O': 30,  # Notes
        }
        
        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width
    
    def _invoice_to_row(self, invoice: Invoice) -> List[Any]:
        """تحويل فاتورة إلى صف Excel"""
        return [
            invoice.invoice_number,
            invoice.invoice_date.isoformat() if invoice.invoice_date else '',
            invoice.vendor.name if invoice.vendor else '',
            invoice.vendor.tax_id if invoice.vendor else '',
            float(invoice.subtotal) if invoice.subtotal else 0,
            float(invoice.total_tax) if invoice.total_tax else 0,
            float(invoice.total_discount) if invoice.total_discount else 0,
            float(invoice.total_amount) if invoice.total_amount else 0,
            invoice.currency.value if invoice.currency else 'SAR',
            len(invoice.line_items),
            invoice.language_detected.value if invoice.language_detected else 'unknown',
            round(invoice.confidence_score, 2) if invoice.confidence_score else 0,
            datetime.now().isoformat(),
            'Processed',
            ''
        ]
    
    def _format_row(self, row_num: int):
        """تنسيق صف"""
        # تنسيق الأرقام
        number_columns = ['E', 'F', 'G', 'H']  # Subtotal, Tax, Discount, Total
        for col in number_columns:
            cell = self.worksheet[f'{col}{row_num}']
            cell.number_format = '#,##0.00'
        
        # محاذاة
        for col_num in range(1, len(self.headers) + 1):
            cell = self.worksheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(vertical="center")
        
        # حدود
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num in range(1, len(self.headers) + 1):
            cell = self.worksheet.cell(row=row_num, column=col_num)
            cell.border = thin_border
    
    def _check_duplicate(self, invoice_number: str) -> bool:
        """التحقق من وجود فاتورة مكررة"""
        if not self.worksheet:
            return False
        
        # البحث في العمود الأول (Invoice Number)
        for row in range(2, self.worksheet.max_row + 1):
            cell_value = self.worksheet.cell(row=row, column=1).value
            if cell_value == invoice_number:
                return True
        
        return False
    
    def get_invoice_status(self, erp_id: str) -> Dict[str, Any]:
        """الحصول على حالة فاتورة"""
        if not self.worksheet:
            self.connect()
        
        # البحث عن الفاتورة
        for row in range(2, self.worksheet.max_row + 1):
            if self.worksheet.cell(row=row, column=1).value == erp_id:
                return {
                    "found": True,
                    "invoice_number": erp_id,
                    "status": self.worksheet.cell(row=row, column=14).value,
                    "row_number": row
                }
        
        return {"found": False}


# ═══════════════════════════════════════════════════
# تسجيل في Factory
# ═══════════════════════════════════════════════════

from .base_connector import ConnectorFactory
ConnectorFactory.register(ConnectorType.EXCEL, ExcelConnector)


__all__ = ['ExcelConnector']